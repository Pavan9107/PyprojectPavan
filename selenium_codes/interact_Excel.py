import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ASUS\\Desktop\\Book1.xlsx")
sheet = book.active

cell = sheet.cell(row=1, column=1)
print(cell.value)

sheet.cell(row=1, column=2).value = 'Updated value'
print(sheet.cell(row=1, column=2).value)



